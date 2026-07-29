"""Honey Tribe LOCAL data builder — Stage 2 stand-in, off-cloud.

Reads the client's context workbook (`context/Honey Tribe (1).xlsx` — the sheet Windsor and the
Colab notebook currently feed) and writes the SAME `honeytribe.json` shape that the live export
job will write. That keeps the three-stage data contract intact while the connectors are being
stood up: the dashboard is built against the real numbers today, and swapping in the live job
later changes nothing downstream.

    context/*.xlsx  ->  build_local.py (HERE)  ->  data/honeytribe.json  ->  dash/dashboard.html

Shape written (matched BY NAME to dashboard.html's DATA.*):
  { client, tagline, location, currency, generated_at, data_through, brand{}, source{},
    shopify{ orders[], lines[], range[] }, meta{ rows[], range[], campaigns[], adsets[], ads[] },
    trend[], seasons[], sessions{}, demographics{} }

The workbook holds live customer PII (emails, names, addresses, phones). NONE of it is emitted:
customers are reduced to a salted hash used only to count distinct people and to classify an
order as first-time vs returning.

Usage:  python clients/client_honeytribe/job/build_local.py
"""
import base64
import collections
import datetime
import hashlib
import json
import os
import re

import categorize

CLIENT = "honeytribe"
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
XLSX = os.environ.get("HONEYTRIBE_XLSX", os.path.join(_ROOT, "context", "Honey Tribe (1).xlsx"))
OUT = os.environ.get("HONEYTRIBE_LOCAL_OUT", os.path.join(_ROOT, "data", "%s.json" % CLIENT))

# Customer ids are hashed, never emitted in the clear. The salt only has to be stable per build.
_SALT = "agora-honeytribe-v1"

# Sizes that are real apparel sizes. The Shopify export's Size column also picks up shipping
# insurance line items (whose "size" is a price), so anything not in this set is dropped.
_SIZES = ["XX-Small", "X-Small", "Small", "Small/Medium", "Medium", "Medium/Large", "Large",
          "X-Large", "XX-Large", "XXX-Large", "OS", "One Size"]

# The seasonal retail guide shown on the Product Analysis tab (client's own planning table).
SEASONS = [
    (1, "January", "Winter", "Winter / New Year / Resort",
     "New-year styling, vacation looks, clearance, fresh wardrobe energy"),
    (2, "February", "Winter", "Valentine / Early Spring",
     "Date-night, feminine looks, early spring styling"),
    (3, "March", "Spring", "Spring / Occasion",
     "Brunch, church, Easter, fresh colors, dresses"),
    (4, "April", "Spring", "Spring / Occasion",
     "Easter, church outfits, brunch, spring events"),
    (5, "May", "Spring", "Spring / Graduation / Mother's Day / Vacation Prep",
     "Dresses, sets, gifting, celebration looks"),
    (6, "June", "Summer", "Summer / Vacation",
     "Shorts, sets, bold prints, travel looks"),
    (7, "July", "Summer", "Summer / Vacation / Early Fall Preview",
     "Warm-weather pieces, denim launches, restocks"),
    (8, "August", "Summer", "Late Summer / Back-to-School / Fall Transition",
     "Transitional styling, denim, versatile pieces"),
    (9, "September", "Fall", "Fall Transition",
     "Denim, layering, statement outfits, event looks"),
    (10, "October", "Fall", "Fall / Event / Pre-Holiday",
     "Statement pieces, richer colors, early holiday prep"),
    (11, "November", "Fall", "Holiday / Black Friday-Cyber Monday",
     "Biggest promo period, gifting, best sellers"),
    (12, "December", "Winter", "Holiday / Gifting / Winter / Resort",
     "Holiday outfits, gifts, vacation/resort pieces"),
]

def _num(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _date(x):
    """Normalise a cell to an ISO date string, or None."""
    if x is None or x == "":
        return None
    if isinstance(x, (datetime.datetime, datetime.date)):
        return x.strftime("%Y-%m-%d")
    s = str(x).strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return "%s-%02d-%02d" % (m.group(3), int(m.group(1)), int(m.group(2)))
    return None


def _hash(email):
    if not email:
        return None
    return hashlib.sha1((_SALT + str(email).strip().lower()).encode("utf-8")).hexdigest()[:12]


def _sheet(wb, name):
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [(str(c).strip() if c is not None else "") for c in next(rows)]
    idx = {h: i for i, h in enumerate(header) if h}
    return idx, [r for r in rows if r and any(c is not None for c in r)]


def _get(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


# --- Shopify --------------------------------------------------------------------------------
def build_shopify(wb):
    """Orders + line items from the Shopify export.

    The Shopify CSV shape puts order-level fields on the FIRST row of an order only; later
    line-item rows repeat the order name with everything else blank. So header fields are
    forward-filled per order name before anything is aggregated.

    `Customer Type` in the sheet is computed per ROW by the upstream notebook, which mislabels
    the 2nd..nth line of a multi-line order as "Returning". It is recomputed here at ORDER
    level: a customer's earliest order is first-time, every later one is returning.
    """
    idx, rows = _sheet(wb, "Output Shopify")

    orders, lines = {}, []
    order_seq = []
    cur = None
    for r in rows:
        name = _get(r, idx, "Name")
        if not name:
            continue
        d = _date(_get(r, idx, "Created at"))
        if d:  # first row of this order — it carries the order header
            cur = {
                "name": name,
                "d": d,
                "sub": round(_num(_get(r, idx, "Subtotal")), 2),
                "tot": round(_num(_get(r, idx, "Total")), 2),
                "disc": round(_num(_get(r, idx, "Discount Amount")), 2),
                "code": (_get(r, idx, "Discount Code") or "") or None,
                "pr": (_get(r, idx, "Billing Province")
                       or _get(r, idx, "Shipping Province") or None),
                "email": _get(r, idx, "Email"),
                "fin": _get(r, idx, "Financial Status"),
                "q": 0,
            }
            if name not in orders:
                orders[name] = cur
                order_seq.append(name)
            else:
                cur = orders[name]
        if cur is None:
            continue

        item = _get(r, idx, "Lineitem name")
        qty = int(_num(_get(r, idx, "Lineitem quantity")))
        price = _num(_get(r, idx, "Lineitem price"))
        product = _get(r, idx, "Product") or (str(item).split(" - ")[0] if item else None)
        size = _get(r, idx, "Size")
        cur["q"] += qty
        if item and not categorize.NON_PRODUCT.search(str(item)):
            lines.append({
                "o": name,
                "p": str(product).strip(),
                "sz": str(size).strip() if size in _SIZES else None,
                "q": qty,
                "amt": round(price * qty, 2),
            })

    # Order-level first-time / returning, by customer, in chronological order.
    seq = sorted(orders.values(), key=lambda o: (o["d"], o["name"]))
    seen = set()
    for o in seq:
        cid = _hash(o["email"])
        o["cid"] = cid
        if cid and cid in seen:
            o["ct"] = "ret"
        else:
            o["ct"] = "first"
            if cid:
                seen.add(cid)

    pos = {o["name"]: i for i, o in enumerate(seq)}
    out_orders = [{"d": o["d"], "sub": o["sub"], "tot": o["tot"], "q": o["q"],
                   "pr": o["pr"], "ct": o["ct"], "cid": o["cid"], "disc": o["disc"]}
                  for o in seq]

    # Decorate each line with its order's date / state / customer type so product data filters
    # on exactly the same controls as sales data. `oi` indexes back into orders[], which lets
    # the dashboard answer "which ORDERS contain this product" for an exact cross-filter.
    out_lines = []
    for ln in lines:
        i = pos.get(ln["o"])
        if i is None:
            continue
        o = seq[i]
        out_lines.append({"d": o["d"], "oi": i, "p": ln["p"], "sz": ln["sz"], "q": ln["q"],
                          "amt": ln["amt"], "pr": o["pr"], "ct": o["ct"]})

    dates = [o["d"] for o in seq if o["d"]]
    return out_orders, out_lines, [min(dates), max(dates)] if dates else [None, None]


# --- Meta -----------------------------------------------------------------------------------
def build_meta(wb):
    """Union of the two Windsor ad exports, de-duplicated on (date, campaign, adset, ad).

    `ArchiveFacebook Ads` is the long history and `Automatic Ads` the recent daily pull; they
    overlap, so the newer sheet wins on a collision.
    """
    seen, rows = {}, []
    for sheet in ("ArchiveFacebook Ads", "Automatic Ads"):
        if sheet not in wb.sheetnames:
            continue
        idx, data = _sheet(wb, sheet)
        for r in data:
            d = _date(_get(r, idx, "date"))
            if not d:
                continue
            camp = (_get(r, idx, "campaign") or "").strip()
            adset = (_get(r, idx, "adset_name") or "").strip()
            ad = (_get(r, idx, "ad_name") or "").strip()
            key = (d, camp, adset, ad)
            seen[key] = {
                "d": d, "camp": camp, "adset": adset, "ad": ad,
                "spend": round(_num(_get(r, idx, "spend")), 2),
                "imps": int(_num(_get(r, idx, "impressions"))),
                "clicks": int(_num(_get(r, idx, "clicks"))),
                "lclk": int(_num(_get(r, idx, "link_clicks"))),
                "reach": int(_num(_get(r, idx, "reach"))),
                "freq": round(_num(_get(r, idx, "frequency")), 4),
                "atc": _num(_get(r, idx, "actions_add_to_cart")),
                "pur": _num(_get(r, idx, "actions_offsite_conversion_fb_pixel_purchase")),
                "rev": round(_num(_get(r, idx, "action_values_omni_purchase")), 2),
            }
    rows = sorted(seen.values(), key=lambda x: (x["d"], x["camp"], x["ad"]))
    dates = [r["d"] for r in rows]
    return {
        "rows": rows,
        "range": [min(dates), max(dates)] if dates else [None, None],
        "campaigns": sorted({r["camp"] for r in rows if r["camp"]}),
        "adsets": sorted({r["adset"] for r in rows if r["adset"]}),
        "ads": sorted({r["ad"] for r in rows if r["ad"]}),
    }


# --- Product trend --------------------------------------------------------------------------
def build_trend(wb):
    idx, data = _sheet(wb, "Honey Tribe Product Trend Analy")
    out = []
    for r in data:
        title = _get(r, idx, "Product title")
        m = _get(r, idx, "Month of year")
        if not title or not m:
            continue
        out.append({
            "t": str(title).strip(),
            "m": int(_num(m)),
            "mn": (_get(r, idx, "Month Name") or "").strip(),
            "se": (_get(r, idx, "Season") or "").strip(),
            "fs": (_get(r, idx, "Fashion Season") or "").strip(),
            "u": int(_num(_get(r, idx, "Net items sold"))),
            "s": round(_num(_get(r, idx, "Total sales")), 2),
            "o": int(_num(_get(r, idx, "Orders"))),
            "cat": (_get(r, idx, "Product Category") or "Uncategorized").strip(),
        })
    return out


def _b64(path, mime):
    try:
        with open(path, "rb") as fh:
            return "data:%s;base64,%s" % (mime, base64.b64encode(fh.read()).decode())
    except OSError as e:
        print("  logo skip: %s" % e)
        return None


def main():
    import openpyxl  # imported lazily so the module is importable without the dep

    print("[%s] reading %s" % (CLIENT, XLSX))
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    orders, lines, srange = build_shopify(wb)
    meta = build_meta(wb)
    trend = build_trend(wb)

    # Learn the taxonomy from the trend sheet, FREEZE it into assets/categories.json (so the
    # live export job classifies identically without the sheet), then apply it.
    model = categorize.learn(trend)
    categorize.save(model)
    classify = categorize.Classifier(model).classify
    missing = collections.Counter()
    for ln in lines:
        ln["cat"] = classify(ln["p"])
        if ln["cat"] == "Uncategorized":
            missing[ln["p"]] += ln["q"]

    data = {
        "client": "Honey Tribe",
        "tagline": "Bold prints. Modern cuts.",
        "location": "United States",
        "currency": "USD",
        "generated_at": datetime.datetime.now(datetime.timezone.utc)
                                 .strftime("%Y-%m-%dT%H:%M:%SZ"),
        "data_through": max(x for x in [srange[1], meta["range"][1]] if x),
        "source": {
            "shopify": "Shopify (midget-giraffe) order export",
            "meta": "Meta (Facebook) Ads via Windsor.ai",
            "mode": "local build from context workbook",
        },
        "brand": {
            "mark": _b64(os.path.join(_HERE, "assets", "honeytribe-mark.png"), "image/png"),
            "agora_logo": _b64(os.path.join(_HERE, "assets", "agora.png"), "image/png"),
        },
        "shopify": {"orders": orders, "lines": lines, "range": srange},
        "meta": meta,
        "trend": trend,
        "seasons": [{"m": m, "mn": mn, "se": se, "fs": fs, "why": why}
                    for (m, mn, se, fs, why) in SEASONS],
        # Not yet connected — see README. The dashboard renders a wired-but-empty state.
        "sessions": {"enabled": False, "rows": [],
                     "note": "Shopify session data needs the ShopifyQL sessions pull "
                             "(RAW_Shopify_Sessions) wired into the export job."},
        # the offline stand-in has no Meta creative data - the gallery renders its wired-empty
        # state, exactly as it would if the creative pull failed
        "creatives": {"enabled": False, "items": [], "error":
                      "Creative previews need the live Windsor pull (job/main.py).",
                      "window": ""},
        "demographics": {"enabled": False, "age_gender": [], "region": [],
                         "note": "Age/gender come from Meta's demographic breakdown pulls; "
                                 "Shopify does not carry them."},
    }

    body = json.dumps(data, separators=(",", ":"))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:  # ASCII-safe JSON, so text mode is fine here
        fh.write(body)

    print("[%s] wrote %s" % (CLIENT, OUT))
    print("   orders   %6d   %s .. %s" % (len(orders), srange[0], srange[1]))
    print("   lines    %6d   (%d uncategorised titles)" % (len(lines), len(missing)))
    print("   meta     %6d   %s .. %s  (%d campaigns, %d ads)"
          % (len(meta["rows"]), meta["range"][0], meta["range"][1],
             len(meta["campaigns"]), len(meta["ads"])))
    print("   trend    %6d" % len(trend))
    print("   size     %6d KB" % (len(body) // 1024))
    if missing:
        print("   top uncategorised: %s" % ", ".join(t for t, _ in missing.most_common(6)))


if __name__ == "__main__":
    main()
