"""MeloYelo LOCAL data builder — Stage 2 stand-in, off-cloud.

Reads the client's context drop (the two workbooks the Looker Studio report is fed from, which
are themselves refreshed by the client's Colab pipelines) and writes `data/meloyelo.json` — the
same shape the future live export job will write. The dashboard is built against real numbers
today; swapping in the live job later changes nothing downstream.

    context/Melo Yelo Unleashed Data (7).xlsx   -> Unleashed sales lines (the reconciled feed:
                                                   revenue = invoices - credits, COGS locked at
                                                   time of sale; see Unleashed_API_Final.ipynb)
                                                   + Lark/Unleashed inventory + production tabs
    context/Customer Data - master (4).xlsx     -> the live CRM (riders, leads, stage history)
    context/Campaign_Monitor_*.ipynb            -> email extract summary (totals only)

The lead-gen pipeline REPLICATES the client-approved Colab logic (cell 17 of the Campaign
Monitor notebook): Warranty Registration rows are excluded from lead counts, requests dedupe by
Unique ID keeping the earliest date, stage events = Stage history + an All-data fallback row per
(id, stage) not already in history, and every count is COUNT-DISTINCT Unique ID per metric.

PII: the CRM holds names, emails, phones, street addresses. NONE of it is emitted — a person is
reduced to their numeric CRM id, dates, stage, agent, source, postcode (-> region) and bike model.

Usage:  py clients/client_MeloYelo/job/build_local.py
"""
import base64
import datetime
import json
import os
import re

CLIENT = "meloyelo"
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
XLSX_SALES = os.environ.get(
    "MELOYELO_SALES_XLSX", os.path.join(_ROOT, "context", "Melo Yelo Unleashed Data (7).xlsx"))
XLSX_CRM = os.environ.get(
    "MELOYELO_CRM_XLSX", os.path.join(_ROOT, "context", "Customer Data - master (4).xlsx"))
NB_EMAIL = os.environ.get(
    "MELOYELO_EMAIL_NB",
    os.path.join(_ROOT, "context", "Campaign_Monitor_Full_Email_Extract (3).ipynb"))
OUT = os.environ.get("MELOYELO_LOCAL_OUT", os.path.join(_ROOT, "data", "%s.json" % CLIENT))

FY_START_MONTH = 4                 # NZ financial year starts 1 April
ANNUAL_BIKE_TARGET = 900           # the client's stated FY target (Looker gauge)
EXCLUDE_SOURCES = {"warranty registration"}   # never counted as lead-gen activity

# Bike unit/revenue counts include the legacy bundle group (verified: FYTD 157 units and the
# prior-FYTD 187 both reproduce Looker's Sales Performance page only with this set).
BIKE_GROUPS = {"Bikes", "Special Bike Package"}
PA_GROUPS = {"Parts", "Accessories"}

# The one bike-model vocabulary (same list the client's notebook uses).
_MODELS = ["SuperLite3", "SuperTrail", "SuperLite", "TownTrail", "Town'n Trail", "Town n Trail",
           "Townee 2.0", "Townee 1.0", "Townee", "Traverse MD", "Traverse", "Ascent3",
           "Ascent MD", "Ascent", "Tranzit MD", "Tranzit", "Roam", "Zoomin", "Flamingo",
           "Kirin", "Rhino"]

# NZ postcode -> region, by the well-known postcode blocks. Approximate on purpose (postcodes
# don't map 1:1 to regions at the edges) and labelled as such on the dashboard.
_PC_REGIONS = [
    (100, 599, "Northland"),      # Whangarei 011x, Kerikeri 023x, Kaitaia 041x, Waipu/Mangawhai 05xx
    (600, 949, "Auckland"),       # North Shore 06xx, west/Helensville 08xx, Warkworth/Orewa 09xx
    (950, 999, "Northland"),      # far-north rural RD codes (e.g. 0981, Whangarei surrounds)
    (1000, 2999, "Auckland"),
    (3000, 3199, "Bay of Plenty"),
    (3200, 3999, "Waikato"),
    (4000, 4099, "Gisborne"),
    (4100, 4299, "Hawke's Bay"),
    (4300, 4399, "Taranaki"),
    (4400, 4599, "Manawatū-Whanganui"),
    (4600, 4699, "Taranaki"),
    (4700, 4999, "Manawatū-Whanganui"),
    (5000, 6999, "Wellington"),
    (7000, 7399, "Nelson-Marlborough"),
    (7400, 7799, "Canterbury"),
    (7800, 7999, "West Coast"),
    (8000, 8999, "Canterbury"),
    (9000, 9599, "Otago"),
    (9600, 9999, "Southland"),
]

# City/town -> region fallback for people with no usable postcode (most warranty imports).
_CITY_REGIONS = {
    "Northland": ["whangarei", "whangārei", "kerikeri", "kaitaia", "dargaville", "kaikohe",
                  "paihia", "mangawhai", "kaiwaka", "waipu", "russell", "kawakawa", "kamo",
                  "tutukaka", "mangonui", "opua", "ruakaka", "one tree point", "hikurangi"],
    "Auckland": ["auckland", "north shore", "manukau", "waitakere", "papakura", "pukekohe",
                 "warkworth", "orewa", "waiuku", "helensville", "albany", "howick", "botany",
                 "takanini", "silverdale", "whangaparaoa", "beachlands", "kumeu", "wellsford",
                 "snells beach", "waiheke", "devonport", "henderson", "titirangi", "remuera",
                 "epsom", "onehunga", "mt eden", "mount eden", "glenfield", "browns bay",
                 "torbay", "red beach", "stanmore bay", "gulf harbour", "half moon bay",
                 "flat bush", "swanson", "massey", "te atatu", "new lynn", "avondale",
                 "mt roskill", "papatoetoe", "mangere", "drury", "clarks beach", "beach haven"],
    "Waikato": ["hamilton", "cambridge", "te awamutu", "taupo", "taupō", "tokoroa", "matamata",
                "morrinsville", "te kuiti", "huntly", "ngaruawahia", "raglan", "otorohanga",
                "putaruru", "turangi", "thames", "whitianga", "coromandel", "whangamata",
                "paeroa", "waihi", "te aroha", "tairua", "pauanui", "cooks beach", "tuakau",
                "te kauwhata", "pokeno", "taumarunui"],
    "Bay of Plenty": ["tauranga", "rotorua", "whakatane", "whakatāne", "mount maunganui",
                      "mt maunganui", "papamoa", "te puke", "kawerau", "opotiki", "katikati",
                      "omokoroa", "waihi beach", "ohope"],
    "Gisborne": ["gisborne"],
    "Hawke's Bay": ["napier", "hastings", "havelock north", "waipukurau", "wairoa", "clive",
                    "taradale", "waipawa"],
    "Taranaki": ["new plymouth", "hawera", "stratford", "waitara", "inglewood", "opunake",
                 "eltham", "bell block", "oakura"],
    "Manawatū-Whanganui": ["palmerston north", "whanganui", "wanganui", "levin", "feilding",
                           "marton", "dannevirke", "foxton", "bulls", "ohakune", "pahiatua",
                           "woodville", "raetihi", "taihape", "ashhurst", "foxton beach",
                           "shannon", "sanson"],
    "Wellington": ["wellington", "lower hutt", "upper hutt", "porirua", "paraparaumu",
                   "kapiti", "waikanae", "masterton", "carterton", "greytown", "featherston",
                   "martinborough", "otaki", "petone", "wainuiomata", "johnsonville",
                   "plimmerton", "pukerua bay", "eastbourne", "paekakariki", "raumati",
                   "tawa", "newlands", "karori", "miramar"],
    "Nelson-Marlborough": ["nelson", "blenheim", "motueka", "richmond", "picton", "takaka",
                           "mapua", "wakefield", "brightwater", "stoke", "havelock", "seddon",
                           "renwick"],
    "West Coast": ["greymouth", "hokitika", "westport", "reefton", "runanga", "franz josef",
                   "karamea"],
    "Canterbury": ["christchurch", "timaru", "ashburton", "rangiora", "kaiapoi", "rolleston",
                   "lincoln", "geraldine", "temuka", "waimate", "amberley", "darfield",
                   "leeston", "oxford", "akaroa", "methven", "fairlie", "twizel",
                   "hanmer springs", "cheviot", "culverden", "kaikoura", "kaikōura", "pegasus",
                   "prebbleton", "west melton", "woodend", "rakaia", "pleasant point",
                   "diamond harbour", "lyttelton", "sumner", "halswell", "riccarton"],
    "Otago": ["dunedin", "queenstown", "wanaka", "wānaka", "oamaru", "alexandra", "cromwell",
              "balclutha", "mosgiel", "milton", "roxburgh", "arrowtown", "clyde", "ranfurly",
              "port chalmers", "brighton", "lawrence", "middlemarch", "omakau", "luggate",
              "hawea", "lake hawea"],
    "Southland": ["invercargill", "gore", "te anau", "winton", "riverton", "bluff", "otautau",
                  "wyndham", "lumsden", "mataura", "stewart island"],
}
_CITY_LOOKUP = {}
for _rg, _cities in _CITY_REGIONS.items():
    for _c in _cities:
        _CITY_LOOKUP[_c] = _rg


def region_of(pc, city=None):
    """Region from a postcode (Excel-float-safe: '981.0' means 0981), else from the city name."""
    if pc not in (None, ""):
        s = str(pc).strip()
        s = re.sub(r"\.0+$", "", s)          # 981.0 -> 981 (a float-mangled 0981)
        s = re.sub(r"\D", "", s)
        if s:
            if len(s) == 3:                  # Excel dropped the leading zero
                s = "0" + s
            n = int(s[:4])
            for lo, hi, name in _PC_REGIONS:
                if lo <= n <= hi:
                    return name
    if city:
        c = re.sub(r"[^a-zāēīōū' ]", " ", str(city).lower())
        c = re.sub(r"\s+", " ", c).strip()
        if c in _CITY_LOOKUP:
            return _CITY_LOOKUP[c]
        for name in _CITY_LOOKUP:           # 'Whangarei, Northland' etc.
            if len(name) > 5 and name in c:
                return _CITY_LOOKUP[name]
    return None


def _num(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(str(x).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


_SHEETS_EPOCH = datetime.datetime(1899, 12, 30)


def _date(x):
    """Normalise a cell to an ISO date string, or None. Accepts datetimes, ISO / NZ day-first
    strings, and Google-Sheets serial numbers (days since 1899-12-30, the UNFORMATTED_VALUE
    shape the live Sheets pull returns)."""
    if x is None or x == "":
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        if 10000 < x < 80000:      # 1927..2119 — a Sheets date serial
            return (_SHEETS_EPOCH + datetime.timedelta(days=float(x))).strftime("%Y-%m-%d")
        return None
    if isinstance(x, (datetime.datetime, datetime.date)):
        return x.strftime("%Y-%m-%d")
    s = str(x).strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:  # NZ day-first
        return "%s-%02d-%02d" % (m.group(3), int(m.group(2)), int(m.group(1)))
    return None


def _dt_minutes(x):
    """Datetime cell -> ISO minute string (for stage-event ordering), or None."""
    if isinstance(x, datetime.datetime):
        return x.strftime("%Y-%m-%d %H:%M")
    if isinstance(x, (int, float)) and not isinstance(x, bool) and 10000 < x < 80000:
        return (_SHEETS_EPOCH + datetime.timedelta(days=float(x))).strftime("%Y-%m-%d %H:%M")
    d = _date(x)
    return (d + " 00:00") if d else None


def _clean_id(v):
    if v is None:
        return ""
    s = str(v).strip()
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def _sheet(wb, name):
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [(str(c).strip() if c is not None else "") for c in next(rows)]
    idx = {h: i for i, h in enumerate(header) if h}
    return idx, [r for r in rows if r and any(c is not None for c in r)]


def _get(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    return row[i]


def _s(v):
    return str(v).strip() if v is not None else ""


def extract_model(product):
    """The client's product-title -> bike-model rule (from Unleashed_API_Final.ipynb)."""
    if not product:
        return ""
    p = str(product).strip()
    low = p.lower()
    for m in _MODELS:
        if m.lower() in low:
            return m
    return ""


# --- Unleashed sales lines --------------------------------------------------------------------
def build_sales(wb):
    idx, rows = _sheet(wb, "Sales Line Data")
    out = []
    for r in rows:
        d = _date(_get(r, idx, "Date"))
        if not d:
            continue
        grp = _s(_get(r, idx, "Product Group")) or "Other"
        model = _s(_get(r, idx, "Bike Model"))
        prod = _s(_get(r, idx, "Product"))
        if not model and grp in BIKE_GROUPS:
            model = extract_model(prod)
        out.append({
            "d": d,
            "o": _s(_get(r, idx, "Order No.")),
            "t": "C" if _s(_get(r, idx, "Doc Type")) == "Credit Note" else "I",
            "ag": _s(_get(r, idx, "Agent")),
            "as": _s(_get(r, idx, "Agent Status")) or "Active",
            "ct": _s(_get(r, idx, "Customer Type")) or "Unknown",
            "g": grp,
            "m": model,
            "p": prod[:80],
            "u": round(_num(_get(r, idx, "Units")), 2),
            "r": round(_num(_get(r, idx, "Revenue")), 2),
            "c": round(_num(_get(r, idx, "COGS")), 2),
        })
    out.sort(key=lambda x: x["d"])
    dates = [x["d"] for x in out]
    return out, [dates[0], dates[-1]] if dates else [None, None]


# --- CRM: riders + lead-gen events (replicates the Colab pipeline) ------------------------------
def build_crm(wb):
    a_idx, a_rows = _sheet(wb, "All data")
    h_idx, h_rows = _sheet(wb, "Stage history")

    people = {}
    for r in a_rows:
        pid = _clean_id(_get(r, a_idx, "Unique ID"))
        d = _date(_get(r, a_idx, "Date added"))
        if not pid or pid in people:
            continue
        people[pid] = {
            "id": pid,
            "d": d,
            "stage": _s(_get(r, a_idx, "Stage")),
            "ag": _s(_get(r, a_idx, "Agent")),
            "src": _s(_get(r, a_idx, "Source")),
            "pc": _s(_get(r, a_idx, "Post Code")),
            "city": _s(_get(r, a_idx, "City")),
            "m": extract_model(_s(_get(r, a_idx, "Model"))) or _s(_get(r, a_idx, "Model"))[:24],
            "stl": _num(_get(r, a_idx, "speed_to_lead_minutes")) or None,
        }

    def excluded(pid):
        p = people.get(pid)
        return bool(p) and p["src"].lower() in EXCLUDE_SOURCES

    # Stage events: Stage history first, then an All-data fallback row for any (id, stage)
    # combination history doesn't cover (exactly the Colab's two-source union).
    events = []          # {id, d (minute), stage, from, ag}
    hist_keys = set()
    for r in h_rows:
        pid = _clean_id(_get(r, h_idx, "Unique ID"))
        to = _s(_get(r, h_idx, "To Stage"))
        d = _dt_minutes(_get(r, h_idx, "Changed At"))
        if not pid or not to or not d:
            continue
        events.append({"id": pid, "d": d, "stage": to,
                       "from": _s(_get(r, h_idx, "From Stage")),
                       "ag": _s(_get(r, h_idx, "Agent")) or (people.get(pid) or {}).get("ag", "")})
        hist_keys.add(pid + "||" + to)
    for p in people.values():
        if p["d"] and p["stage"] and not excluded(p["id"]) \
                and (p["id"] + "||" + p["stage"]) not in hist_keys:
            events.append({"id": p["id"], "d": p["d"] + " 00:00", "stage": p["stage"],
                           "from": "", "ag": p["ag"]})
    events.sort(key=lambda e: e["d"])

    # Riders: everyone whose CURRENT stage is MY Customer. Joined = the first recorded
    # transition INTO MY Customer, else the date they entered the CRM. This is the fix for the
    # Looker "Riders Over Time" chart, which plotted rider counts against customer NAMES.
    first_my = {}
    for e in events:
        if e["stage"] == "MY Customer" and e["id"] not in first_my:
            first_my[e["id"]] = e["d"][:10]
    riders = []
    for p in people.values():
        if p["stage"] != "MY Customer":
            continue
        joined = first_my.get(p["id"]) or p["d"]
        riders.append({"id": p["id"], "d": joined, "ag": p["ag"],
                       "src": p["src"] or "Unattributed",
                       "rg": region_of(p["pc"], p["city"]), "city": p["city"], "m": p["m"]})
    riders.sort(key=lambda x: x["d"] or "")

    # Lead-gen unified events: Test Ride Request per person (earliest date, warranty excluded)
    # + every stage event for non-warranty people. Metric counts are COUNT-DISTINCT ids.
    unified = []
    for p in people.values():
        if p["d"] and not excluded(p["id"]):
            unified.append({"id": p["id"], "d": p["d"], "m": "Test Ride Request",
                            "ag": p["ag"], "src": p["src"] or "",
                            "rg": region_of(p["pc"], p["city"])})
    for e in events:
        if excluded(e["id"]):
            continue
        p = people.get(e["id"]) or {}
        unified.append({"id": e["id"], "d": e["d"][:10], "m": e["stage"],
                        "ag": e["ag"] or p.get("ag", ""), "src": p.get("src", "") or "",
                        "rg": region_of(p.get("pc"), p.get("city"))})
    unified.sort(key=lambda x: x["d"])

    # Requests (one row per lead, deduped, warranty excluded) power the speed-to-lead view.
    requests = []
    for p in people.values():
        if not p["d"] or excluded(p["id"]):
            continue
        requests.append({"id": p["id"], "d": p["d"], "stage": p["stage"], "ag": p["ag"],
                         "src": p["src"] or "Unattributed", "stl": p["stl"],
                         "rg": region_of(p["pc"], p["city"]), "city": p["city"]})
    requests.sort(key=lambda x: x["d"])

    dates = [p["d"] for p in people.values() if p["d"]]
    return {
        "riders": riders,
        "requests": requests,
        "events": unified,
        "range": [min(dates), max(dates)] if dates else [None, None],
    }


# --- Inventory + production (Unleashed stock join + the Lark base) ------------------------------
_SIZES = {"Small": "Small", "SM": "Small", "Large": "Large", "LG": "Large", "Medium": "Medium",
          "MD": "Medium", "Step-Thru": "Step-Thru", "Step Thru": "Step-Thru"}


def _variant_parts(variant):
    """'Ascent3 Small Moonlight Blue' -> (model, size, colour)."""
    v = _s(variant)
    model = extract_model(v) or (v.split(",")[0].strip() if v else "")
    rest = v
    if model:
        i = v.lower().find(model.lower())
        rest = (v[:i] + v[i + len(model):]).replace(",", " ").strip()
    size = ""
    for k, s in _SIZES.items():
        if re.search(r"\b" + re.escape(k) + r"\b", rest, re.I):
            size = s
            rest = re.sub(r"\b" + re.escape(k) + r"\b", "", rest, flags=re.I).strip()
            break
    colour = re.sub(r"\s{2,}", " ", rest).strip(" ,-")
    return model or v, size, colour


def build_inventory(wb):
    idx, rows = _sheet(wb, "Available Now")
    avail = []
    for r in rows:
        v = _s(_get(r, idx, "Variant"))
        if not v:
            continue
        model, size, colour = _variant_parts(v)
        avail.append({"v": v, "code": _s(_get(r, idx, "Code")),
                      "m": model, "sz": size, "col": colour,
                      "q": int(_num(_get(r, idx, "Available"))),
                      "st": _s(_get(r, idx, "Status"))})
    avail.sort(key=lambda x: -x["q"])

    idx, rows = _sheet(wb, "On The Way")
    onway = []
    for r in rows:
        pm = _s(_get(r, idx, "Product / Model"))
        if not pm:
            continue
        onway.append({
            "m": pm,
            "q": int(_num(_get(r, idx, "Quantity"))),
            "stage": _s(_get(r, idx, "Production stage")),
            "prod": _date(_get(r, idx, "Estimated Production Date")),
            "ship": _date(_get(r, idx, "Estimated Shipping Date")),
            "tl": _s(_get(r, idx, "Timeliness")),
            "sched": _s(_get(r, idx, "Schedule status")),
            "act": _s(_get(r, idx, "Next Action")),
            "po": _s(_get(r, idx, "PO Number")),
            "upd": _date(_get(r, idx, "Last updated")),
        })
    onway.sort(key=lambda x: x["ship"] or "9999")
    return {"available": avail, "onway": onway}


def build_meta_stamp(wb):
    try:
        idx, rows = _sheet(wb, "Dashboard Meta")
        if rows:
            return {"label": _s(_get(rows[0], idx, "Update Label")),
                    "refreshed": _dt_minutes(_get(rows[0], idx, "Last Refresh NZ")),
                    "tz": _s(_get(rows[0], idx, "Timezone"))}
    except KeyError:
        pass
    return None


# --- Campaign Monitor: extract-run summary from the notebook's own outputs ----------------------
def build_email():
    out = {"enabled": False, "campaigns": [], "totals": None,
           "note": "Full per-campaign engagement lands when the Campaign Monitor API pull is "
                   "wired into the export job (the client's extract notebook already proves "
                   "the pull)."}
    try:
        with open(NB_EMAIL, encoding="utf-8") as fh:
            nb = json.load(fh)
    except (OSError, ValueError):
        return out
    text = []
    for c in nb.get("cells", []):
        for o in c.get("outputs", []):
            if "text" in o:
                text.append("".join(o["text"]))
            data = o.get("data", {})
            if "text/plain" in data:
                text.append("".join(data["text/plain"]))
    blob = "\n".join(text)

    m = re.search(r"Starting pull of (\d+) campaigns", blob)
    n_camp = int(m.group(1)) if m else None
    tm = re.search(r"Final totals:\s*\nSends:\s*([\d,]+)\s*\nOpens:\s*([\d,]+)\s*\nClicks:\s*([\d,]+)", blob)
    if not tm:
        tm = re.search(r"Sends:\s*([\d,]+)[\s\S]{0,80}?Opens:\s*([\d,]+)[\s\S]{0,80}?Clicks:\s*([\d,]+)", blob)
    if tm:
        # unique_opens is None (not 0): the notebook's printed totals carry only
        # Sends/Opens/Clicks, and the live pull_email() shape has the key — the
        # two builders must emit the same shape (see CLAUDE.md).
        out["totals"] = {"campaigns": n_camp,
                         "sends": int(tm.group(1).replace(",", "")),
                         "opens": int(tm.group(2).replace(",", "")),
                         "clicks": int(tm.group(3).replace(",", "")),
                         "unique_opens": None}
    seen = {}
    for m in re.finditer(r"\[START\] (.{1,60}?)\s+\(expected ([\d,?]+) recipients\)", blob):
        name = m.group(1).strip()
        try:
            rec = int(m.group(2).replace(",", ""))
        except ValueError:
            continue
        seen[name] = max(rec, seen.get(name, 0))
    camps = sorted(seen.items(), key=lambda kv: -kv[1])
    # date/uopens/uclicks exist in the live pull_email() rows; the notebook's
    # [START] lines carry neither, so they ship as empty/None for shape parity.
    out["campaigns"] = [{"name": k, "date": "", "recipients": v,
                         "uopens": None, "uclicks": None} for k, v in camps]
    dm = re.findall(r"(\d{4}-\d{2}-\d{2}) \d{2}:\d{2}:\d{2}", blob)
    if dm:
        out["extracted_through"] = max(dm)
    return out


def _b64(path, mime):
    try:
        with open(path, "rb") as fh:
            return "data:%s;base64,%s" % (mime, base64.b64encode(fh.read()).decode())
    except OSError as e:
        print("  logo skip: %s" % e)
        return None


def main():
    import openpyxl  # lazy so the module imports without the dep

    print("[%s] reading %s" % (CLIENT, os.path.basename(XLSX_SALES)))
    wb_sales = openpyxl.load_workbook(XLSX_SALES, read_only=True, data_only=True)
    sales, srange = build_sales(wb_sales)
    inv = build_inventory(wb_sales)
    stamp = build_meta_stamp(wb_sales)
    wb_sales.close()

    print("[%s] reading %s" % (CLIENT, os.path.basename(XLSX_CRM)))
    wb_crm = openpyxl.load_workbook(XLSX_CRM, read_only=True, data_only=True)
    crm = build_crm(wb_crm)
    wb_crm.close()

    email = build_email()

    data = {
        "client": "MeloYelo",
        "tagline": "just mad about E-BIKES",
        "location": "New Zealand",
        "currency": "NZD",
        "generated_at": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "data_through": max(x for x in [srange[1], crm["range"][1]] if x),
        "fy_start_month": FY_START_MONTH,
        "targets": {"fy_bikes": ANNUAL_BIKE_TARGET},
        "source": {
            "sales": "Unleashed via the client's reconciled Sales Line feed "
                     "(revenue = invoices - credits, COGS at time of sale)",
            "crm": "Customer Data master sheet (the live CRM)",
            "inventory": "Unleashed stock join + Lark factory-production base",
            "mode": "local build from context workbooks",
        },
        "brand": {
            "mark": _b64(os.path.join(_HERE, "assets", "meloyelo-mark.svg"), "image/svg+xml"),
            "agora_logo": _b64(os.path.join(_HERE, "assets", "agora.png"), "image/png"),
        },
        "stamp": stamp,
        "sales": {"lines": sales, "range": srange},
        "crm": crm,
        "inventory": inv,
        "email": email,
        # Wired, awaiting connectors — the dashboard renders these sections as designed empty
        # states that light up when the export job fills them (Windsor `all` connector for Meta
        # and Google Ads; the GA4 connector for site analytics).
        "meta_ads": {"enabled": False, "rows": [],
                     "note": "Meta per-ad/day rows via Windsor.ai (spend, impressions, reach, "
                             "link clicks, leads, frequency)."},
        "google_ads": {"enabled": False, "rows": [],
                       "note": "Google Ads account 668-008-6591 via Windsor.ai (campaign/day: "
                               "cost, impressions, clicks, conversions)."},
        "ga4": {"enabled": False, "rows": [],
                "note": "GA4 via Windsor.ai (date/channel: sessions, engaged sessions, active "
                        "users, engagement rate)."},
    }

    body = json.dumps(data, separators=(",", ":"))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "wb") as fh:      # BYTES on purpose — never text-mode JSON on Windows
        fh.write(body.encode("utf-8"))

    n_riders = len(crm["riders"])
    print("[%s] wrote %s" % (CLIENT, OUT))
    print("   sales lines %6d   %s .. %s" % (len(sales), srange[0], srange[1]))
    print("   riders      %6d   requests %d   events %d   (CRM %s .. %s)"
          % (n_riders, len(crm["requests"]), len(crm["events"]),
             crm["range"][0], crm["range"][1]))
    print("   inventory   %6d variants   %d production orders"
          % (len(inv["available"]), len(inv["onway"])))
    if email.get("totals"):
        t = email["totals"]
        print("   email       %s campaigns, %s sends, %s opens, %s clicks (extract summary)"
              % (t.get("campaigns"), t.get("sends"), t.get("opens"), t.get("clicks")))
    print("   size        %6d KB" % (len(body) // 1024))

    # Self-check against the client's own QA table logic (COUNT-DISTINCT ids per metric).
    ev = crm["events"]
    for metric in ("Test Ride Request", "Test Ride Booked", "Test Ride Completed"):
        ids = {e["id"] for e in ev if e["m"] == metric}
        print("   QA all-time %-22s %d" % (metric, len(ids)))


if __name__ == "__main__":
    main()
