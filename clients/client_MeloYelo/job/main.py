"""MeloYelo LIVE data builder — pulls the client's real APIs and writes `data/meloyelo.json`.

Same JSON shape as `build_local.py` (the offline stand-in) — the dashboard cannot tell them
apart. What is live vs snapshot, and why:

  LIVE      Unleashed sales lines   Invoices + CreditNotes + SalesOrders + Products + Customers,
                                    the client's reconciled rules (Unleashed_API_Final.ipynb):
                                    revenue = invoices - credits (tax-excl), COGS locked at time
                                    of sale, credit units subtract only when ReturnToStock.
  LIVE      Unleashed stock         Products (Bikes group, current-range prefixes) join
                                    StockOnHand — the "Available now" panel.
  LIVE      Campaign Monitor        campaign list + per-campaign summary (unique opens/clicks),
                                    which the old Looker report never had.
  LIVE*     CRM (riders + leads)    when `CRM_SHEET_ID` is set AND the sheet is shared with the
                                    job's service account (Viewer), the master sheet is read via
                                    the Sheets API (scoped token minted through iamcredentials
                                    self-impersonation — no key file). Fallback: the dropped
                                    `Customer Data - master*.xlsx`, then the previous
                                    publication's CRM section (carry-forward).
  LIVE*     Lark "On the way"       when `LARK_APP_ID`/`LARK_APP_SECRET`/`LARK_APP_TOKEN`/
                                    `LARK_TABLE_ID` are set and a refresh token is available
                                    (seed via `LARK_REFRESH_SEED`; thereafter it SELF-ROTATES —
                                    each run stores the newest token in the bucket, so the
                                    client authorizes ONCE, not every run). Fallback: the
                                    workbook tab, then the previous publication.

Credentials: environment variables first (UNLEASHED_API_ID / UNLEASHED_API_KEY /
CAMPAIGN_MONITOR_API_KEY / CAMPAIGN_MONITOR_CLIENT_ID) — that is how the deployed job will run,
mounted from Secret Manager. Off-cloud, absent env vars, they are read from the notebooks in
`context/` (git-ignored) so a developer can run a real pull with zero setup. Keys are never
printed and never written to the output.

Usage:  py clients/client_MeloYelo/job/main.py
"""
import base64
import concurrent.futures
import datetime
import hashlib
import hmac
import json
import os
import re
import sys
import urllib.parse
import urllib.request

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)
import build_local as bl  # noqa: E402  — the shared transforms + xlsx fallbacks

UNLEASHED_BASE = "https://api.unleashedsoftware.com"
CM_BASE = "https://api.createsend.com/api/v3.3"
LARK_BASE = "https://open.larksuite.com/open-apis"
FULL_START = "2021-04-01"
CURRENT_PREFIXES = ("AS3", "SL3", "ST3", "TNT3", "Z1")   # Greg's current range (Lark notebook)

NZ_TZ = datetime.timezone(datetime.timedelta(hours=12))

# Cloud publication — set on the Cloud Run job; absent locally (writes data/meloyelo.json).
GCS_BUCKET = os.environ.get("GCS_BUCKET", "")
DATA_OBJECT = os.environ.get("DATA_OBJECT", "meloyelo.json")
GCP_PROJECT = os.environ.get("GCP_PROJECT", "agora-data-driven")
CRM_SHEET_ID = os.environ.get("CRM_SHEET_ID", "")
LARK_TOKEN_OBJECT = "lark_token.json"      # rotating refresh token, private, lives in the bucket


def _gcs_bucket():
    from google.cloud import storage   # lazy — local runs don't need it
    return storage.Client(project=GCP_PROJECT).bucket(GCS_BUCKET)


def load_previous():
    """The JSON this job published last time, or None (carry-forward for snapshot sources)."""
    try:
        if GCS_BUCKET:
            blob = _gcs_bucket().blob(DATA_OBJECT)
            if not blob.exists():
                return None
            return json.loads(blob.download_as_bytes().decode("utf-8"))
        if os.path.exists(bl.OUT):
            with open(bl.OUT, encoding="utf-8") as fh:
                return json.load(fh)
    except Exception as e:  # noqa: BLE001 — no previous publication is a normal first run
        print("  previous publication unreadable (%s) — starting fresh" % str(e)[:120])
    return None


# ------------------------------------------------ Google Sheets (CRM) via the runtime SA
def _metadata_token():
    req = urllib.request.Request(
        "http://metadata.google.internal/computeMetadata/v1/instance/service-accounts/default/token",
        headers={"Metadata-Flavor": "Google"})
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.load(r)["access_token"]


def _sa_email():
    req = urllib.request.Request(
        "http://metadata.google.internal/computeMetadata/v1/instance/service-accounts/default/email",
        headers={"Metadata-Flavor": "Google"})
    with urllib.request.urlopen(req, timeout=10) as r:
        return r.read().decode().strip()


def _sheets_token():
    """A spreadsheets.readonly-scoped token for the runtime SA.

    Cloud Run metadata tokens carry only the cloud-platform scope, which the Sheets API
    rejects — so mint a scoped token via iamcredentials self-impersonation (the same keyless
    pattern the Atrium signed-upload uses). Needs roles/iam.serviceAccountTokenCreator on the
    SA itself; the deploy script grants it. Off-cloud, GOOGLE_SHEETS_TOKEN lets a dev test."""
    dev = os.environ.get("GOOGLE_SHEETS_TOKEN")
    if dev:
        return dev
    body = json.dumps({"scope": ["https://www.googleapis.com/auth/spreadsheets.readonly"]}).encode()
    req = urllib.request.Request(
        "https://iamcredentials.googleapis.com/v1/projects/-/serviceAccounts/%s:generateAccessToken"
        % _sa_email(),
        data=body, method="POST",
        headers={"Authorization": "Bearer " + _metadata_token(),
                 "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)["accessToken"]


class _SheetWS:
    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, values_only=True):  # noqa: ARG002 — openpyxl-shaped
        for r in self._rows:
            yield tuple(r)


class _SheetWB(dict):
    """Just enough of the openpyxl workbook surface for bl._sheet()/build_crm()."""


def pull_crm_sheet():
    """LIVE CRM from the master Google Sheet, or None if unconfigured/unshared."""
    if not CRM_SHEET_ID:
        return None
    try:
        token = _sheets_token()
        wb = _SheetWB()
        for tab in ("All data", "Stage history"):
            url = ("https://sheets.googleapis.com/v4/spreadsheets/%s/values/%s"
                   "?majorDimension=ROWS&valueRenderOption=UNFORMATTED_VALUE"
                   "&dateTimeRenderOption=SERIAL_NUMBER"
                   % (CRM_SHEET_ID, urllib.parse.quote("'" + tab + "'")))
            req = urllib.request.Request(url, headers={"Authorization": "Bearer " + token})
            with urllib.request.urlopen(req, timeout=120) as r:
                rows = json.load(r).get("values", [])
            if not rows:
                raise RuntimeError("tab '%s' came back empty" % tab)
            wb[tab] = _SheetWS(rows)
            print("  [crm] sheet tab '%s': %d rows" % (tab, len(rows) - 1), flush=True)
        return bl.build_crm(wb)
    except Exception as e:  # noqa: BLE001 — an unshared sheet degrades, never crashes
        print("  [crm] live sheet pull unavailable (%s) — falling back" % str(e)[:160], flush=True)
        return None


# --------------------------------------------------------- Lark base (production orders)
def _lark_state_load():
    seed = os.environ.get("LARK_REFRESH_SEED", "")
    try:
        if GCS_BUCKET:
            blob = _gcs_bucket().blob(LARK_TOKEN_OBJECT)
            if blob.exists():
                return json.loads(blob.download_as_bytes().decode("utf-8")).get("refresh_token") or seed
    except Exception as e:  # noqa: BLE001
        print("  [lark] token state unreadable (%s)" % str(e)[:120], flush=True)
    return seed


def _lark_state_save(refresh_token):
    try:
        if GCS_BUCKET:
            _gcs_bucket().blob(LARK_TOKEN_OBJECT).upload_from_string(
                json.dumps({"refresh_token": refresh_token}).encode("utf-8"),
                content_type="application/json")
    except Exception as e:  # noqa: BLE001 — losing the rotation is bad but not fatal today
        print("  [lark] could not persist rotated token: %s" % str(e)[:120], flush=True)


def _lark_post(path, payload, bearer=None):
    req = urllib.request.Request(LARK_BASE + path, data=json.dumps(payload).encode(),
                                 method="POST", headers={"Content-Type": "application/json"})
    if bearer:
        req.add_header("Authorization", "Bearer " + bearer)
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.load(r)


def pull_lark():
    """LIVE production orders from the Lark base, or None. The refresh token rotates on every
    use; the newest one is stored back to the bucket so the client authorizes exactly ONCE."""
    app_id = os.environ.get("LARK_APP_ID", "")
    app_secret = os.environ.get("LARK_APP_SECRET", "")
    app_token = os.environ.get("LARK_APP_TOKEN", "")
    table_id = os.environ.get("LARK_TABLE_ID", "")
    refresh = _lark_state_load()
    if not (app_id and app_secret and app_token and table_id and refresh):
        return None
    try:
        app_tok = _lark_post("/auth/v3/app_access_token/internal",
                             {"app_id": app_id, "app_secret": app_secret})["app_access_token"]
        t = _lark_post("/authen/v1/refresh_access_token",
                       {"grant_type": "refresh_token", "refresh_token": refresh}, bearer=app_tok)
        if t.get("code") != 0 or "data" not in t:
            print("  [lark] refresh failed: %s — re-seed LARK_REFRESH_SEED" % str(t)[:160], flush=True)
            return None
        user_token = t["data"]["access_token"]
        _lark_state_save(t["data"]["refresh_token"])

        rows, page = [], None
        while True:
            qs = "page_size=100" + (("&page_token=" + urllib.parse.quote(page)) if page else "")
            req = urllib.request.Request(
                "%s/bitable/v1/apps/%s/tables/%s/records?%s" % (LARK_BASE, app_token, table_id, qs),
                headers={"Authorization": "Bearer " + user_token})
            with urllib.request.urlopen(req, timeout=60) as r:
                d = json.load(r)["data"]
            rows += d.get("items", [])
            if d.get("has_more"):
                page = d.get("page_token")
            else:
                break

        def fdate(v):
            if isinstance(v, (int, float)) and v > 1e11:
                return datetime.datetime.fromtimestamp(v / 1000, NZ_TZ).strftime("%Y-%m-%d")
            return bl._date(v)

        today = datetime.datetime.now(NZ_TZ).strftime("%Y-%m-%d")
        onway = []
        for rec in rows:
            f = rec.get("fields", {})
            pm = f.get("Product / Model")
            if not pm:
                continue
            ship = fdate(f.get("Estimated Shipping Date"))
            stage = str(f.get("Production stage") or "")
            late = bool(ship) and ship < today and "ship" not in stage.lower()
            sched = str(f.get("Schedule status") or "").replace("\U0001F7E2", "").strip()
            onway.append({
                "m": str(pm), "q": int(bl._num(f.get("Quantity"))), "stage": stage,
                "prod": fdate(f.get("Estimated Production Date")), "ship": ship,
                "tl": "Overdue" if late else "On track", "sched": sched,
                "act": str(f.get("Next Action") or ""), "po": str(f.get("PO Number") or ""),
                "upd": fdate(f.get("Last updated")),
            })
        onway.sort(key=lambda x: x["ship"] or "9999")
        print("  [lark] LIVE: %d production orders" % len(onway), flush=True)
        return onway
    except Exception as e:  # noqa: BLE001 — Lark being down never kills the sales publish
        print("  [lark] live pull failed (%s) — falling back" % str(e)[:160], flush=True)
        return None


# ------------------------------------------------------------------- credentials (env first)
def _from_notebook(path, pattern):
    try:
        nb = json.load(open(path, encoding="utf-8"))
    except (OSError, ValueError):
        return None
    src = "".join("".join(c.get("source", [])) for c in nb.get("cells", [])
                  if c.get("cell_type") == "code")
    m = re.search(pattern, src)
    return m.group(1) if m else None


def creds():
    root = os.path.dirname(_HERE)
    lark_nb = os.path.join(root, "context", "Lark_Unleashed_Dash.ipynb")
    cm_nb = os.path.join(root, "context", "Campaign_Monitor_Full_Email_Extract (3).ipynb")
    c = {
        "u_id": os.environ.get("UNLEASHED_API_ID")
                or _from_notebook(lark_nb, r'UNLEASHED_API_ID\s*=\s*"([^"]+)"'),
        "u_key": os.environ.get("UNLEASHED_API_KEY")
                 or _from_notebook(lark_nb, r'UNLEASHED_API_KEY\s*=\s*"([^"]+)"'),
        "cm_key": os.environ.get("CAMPAIGN_MONITOR_API_KEY")
                  or _from_notebook(cm_nb, r'API_KEY\s*=\s*"([^"]+)"'),
        "cm_client": os.environ.get("CAMPAIGN_MONITOR_CLIENT_ID")
                     or _from_notebook(cm_nb, r'CLIENT_ID\s*=\s*"([^"]+)"'),
        "w_key": os.environ.get("WINDSOR_API_KEY") or _windsor_key_local(),
    }
    return c


def _windsor_key_local():
    """Off-cloud convenience: read the Windsor key from Secret Manager as the operator.
    Absent gcloud/permissions this returns None and the Meta pull degrades gracefully."""
    try:
        import subprocess
        r = subprocess.run(
            "gcloud secrets versions access latest --secret meloyelo-windsor-key "
            "--project agora-data-driven --account info@agoradatadriven.com",
            capture_output=True, text=True, shell=True, timeout=30)
        key = (r.stdout or "").strip()
        return key if len(key) > 10 else None
    except Exception:  # noqa: BLE001 — no gcloud locally is normal
        return None


# ----------------------------------------------------------------------- Unleashed plumbing
def unleashed_get(c, endpoint, query=""):
    sig = base64.b64encode(hmac.new(c["u_key"].encode(), query.encode(),
                                    hashlib.sha256).digest()).decode()
    url = UNLEASHED_BASE + "/" + endpoint + (("?" + query) if query else "")
    req = urllib.request.Request(url, headers={
        "Accept": "application/json", "Content-Type": "application/json",
        "api-auth-id": c["u_id"], "api-auth-signature": sig,
        "client-type": "meloyelo/dashboard"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return json.load(r)


def unleashed_all(c, endpoint, params=None, page_size=1000):
    params = dict(params or {})
    params["pageSize"] = page_size
    items, page = [], 1
    while True:
        qs = urllib.parse.urlencode(params)
        d = unleashed_get(c, "%s/Page/%d" % (endpoint, page), qs)
        batch = d.get("Items", [])
        items.extend(batch)
        pages = d.get("Pagination", {}).get("NumberOfPages") or 1
        print("  %s page %d/%d (%d rows)" % (endpoint, page, pages, len(batch)), flush=True)
        if not batch or page >= pages:
            break
        page += 1
    return items


def _ms_date(v):
    """Unleashed /Date(1785177923514)/ -> ISO date (NZ time)."""
    if v in (None, ""):
        return None
    m = re.search(r"/Date\((\d+)", str(v))
    if m:
        return datetime.datetime.fromtimestamp(int(m.group(1)) / 1000, NZ_TZ).strftime("%Y-%m-%d")
    return bl._date(v)


def _gn(obj, path, default=""):
    cur = obj
    for k in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
        if cur is None:
            return default
    return cur


def _num(v):
    try:
        return 0.0 if v in (None, "") else float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _find_lines(record, keys):
    for k in keys:
        v = record.get(k)
        if not v:
            continue
        if isinstance(v, list):
            return v
        if isinstance(v, dict):
            for vv in v.values():
                if isinstance(vv, list):
                    return vv
            return [v]
    return []


# ------------------------------------------------- the client's agent mapping (notebook-exact)
AGENT_MAP = {
    "D.M. George Limited": "Don George",
    "JP & WJ Smith": "Justin Smith",
    "Stew & Sharon Hickford": "Stew Hickford",
    "E-Bike Taranaki Limited": "Shane Blackbourn",
    "Intrinsico Ltd": "Dean Leuschke",
    "Chadsmoor Limited": "Jeff Cook",
    "Velo South Limited": "Shaun Howarth",
    "EBikes & Mobility": "Ebikes & Mobility",
    "E-Mega Store Limited": "E-Mega Store",
    "Simply Furniture t/a eMegastore": "E-Mega Store",
    "Suncoast eBikes Limited": "Suncoast eBikes",
    "Cyclepro Bike Shop": "Cyclepro Bike Shop",
    "EBikes Wanganui": "EBikes Wanganui",
    "Eastern Bikes Limited": "Eastern Bikes",
    "Get Kids Racing": "Get Kids Racing",
    "Workride Limited": "Workride",
    "Lifetime Apprentice Limited - Peter Cajes": "Peter Cajes",
    "Oric Holdings Limited": "Oric Holdings",
}
INACTIVE_AGENTS = {
    "E-Mega Store", "Suncoast eBikes", "Cyclepro Bike Shop", "EBikes Wanganui",
    "Eastern Bikes", "Get Kids Racing", "Workride", "Peter Cajes", "Oric Holdings",
}


def map_agent(customer, cust_type):
    if customer in AGENT_MAP:
        return AGENT_MAP[customer]
    if cust_type == "Website Customer":
        return "Direct Sale"
    if cust_type == "Admin/Sales":
        return "MeloYelo Internal"
    return customer


def agent_status(agent):
    if agent in INACTIVE_AGENTS:
        return "Inactive"
    if agent in ("Direct Sale", "MeloYelo Internal"):
        return agent
    return "Active"


# ----------------------------------------------------------------------- LIVE: sales lines
def pull_sales(c):
    today = datetime.datetime.now(NZ_TZ).strftime("%Y-%m-%d")
    dates = {"startDate": FULL_START, "endDate": today}
    print("[live] Unleashed pull %s -> %s" % (FULL_START, today), flush=True)
    invoices = unleashed_all(c, "Invoices", dates)
    credits = unleashed_all(c, "CreditNotes", dates)
    orders = unleashed_all(c, "SalesOrders", dates)
    products = unleashed_all(c, "Products", {"includeObsolete": "true"})
    customers = unleashed_all(c, "Customers", {"includeObsolete": "true"})

    prod_group = {}
    for p in products:
        code = p.get("ProductCode", "")
        if not code:
            continue
        grp = _gn(p, ["ProductGroup", "GroupName"]) or _gn(p, ["ProductGroup", "Name"])
        if isinstance(p.get("ProductGroup"), str):
            grp = p["ProductGroup"]
        prod_group[code] = grp

    cust_type = {}
    for cu in customers:
        code = cu.get("CustomerCode", "")
        ct = cu.get("CustomerType", "")
        if isinstance(ct, dict):
            ct = ct.get("TypeName", "") or ct.get("Name", "") or ""
        if code:
            cust_type[code] = ct

    cogs = {}
    for o in orders:
        order_no = o.get("OrderNumber", "")
        if not order_no:
            continue
        lines = o.get("SalesOrderLines", [])
        if isinstance(lines, dict):
            lines = lines.get("SalesOrderLine", [])
            if isinstance(lines, dict):
                lines = [lines]
        if not isinstance(lines, list):
            lines = []
        for ln in lines:
            code = _gn(ln, ["Product", "ProductCode"])
            cost = ln.get("AverageLandedPriceAtTimeOfSale")
            if code and cost not in (None, ""):
                cogs[(order_no, code)] = _num(cost)

    out = []

    def emit(d, order_no, doc_type, cust, ctype, code, desc, units, rev, unit_cost):
        grp = (prod_group.get(code) or "").strip() or "Other"
        agent = map_agent(cust, ctype)
        model = bl.extract_model(desc) if grp in bl.BIKE_GROUPS else ""
        out.append({
            "d": d, "o": order_no, "t": doc_type,
            "ag": agent, "as": agent_status(agent),
            "ct": ctype or "Unknown", "g": grp, "m": model, "p": (desc or "")[:80],
            "u": round(units, 2), "r": round(rev, 2), "c": round(units * unit_cost, 2),
        })

    for doc in invoices:
        d = _ms_date(doc.get("InvoiceDate") or doc.get("CreatedOn"))
        if not d:
            continue
        order_no = doc.get("OrderNumber", "")
        cust = _gn(doc, ["Customer", "CustomerName"])
        ctype = cust_type.get(_gn(doc, ["Customer", "CustomerCode"]), "")
        for ln in _find_lines(doc, ["InvoiceLines", "SalesInvoiceLines", "Lines"]):
            if ln.get("LineType") not in (None, "", "Normal"):
                continue
            code = _gn(ln, ["Product", "ProductCode"])
            desc = _gn(ln, ["Product", "ProductDescription"]) or ln.get("LineDescription", "")
            q = ln.get("InvoiceQuantity")
            if q is None:
                q = ln.get("OrderQuantity") or 0
            emit(d, order_no, "I", cust, ctype, code, desc,
                 _num(q), _num(ln.get("LineTotal")), cogs.get((order_no, code), 0.0))

    for doc in credits:
        d = _ms_date(doc.get("CreditDate") or doc.get("CreatedOn"))
        if not d:
            continue
        order_no = _gn(doc, ["SalesOrder", "OrderNumber"]) or doc.get("OrderNumber", "")
        cust = _gn(doc, ["Customer", "CustomerName"])
        ctype = cust_type.get(_gn(doc, ["Customer", "CustomerCode"]), "")
        for ln in _find_lines(doc, ["CreditLines", "CreditNoteLines", "SalesCreditNoteLines", "Lines"]):
            code = _gn(ln, ["Product", "ProductCode"])
            desc = _gn(ln, ["Product", "ProductDescription"]) or ln.get("LineDescription", "")
            qty = _num(ln.get("CreditQuantity") or ln.get("OrderQuantity") or 0)
            unit_cost = _num(ln.get("AverageLandedPriceAtTimeOfSale") or 0)
            grp = (prod_group.get(code) or "").strip() or "Other"
            agent = map_agent(cust, ctype)
            out.append({
                "d": d, "o": order_no, "t": "C",
                "ag": agent, "as": agent_status(agent),
                "ct": ctype or "Unknown", "g": grp,
                "m": bl.extract_model(desc) if grp in bl.BIKE_GROUPS else "",
                "p": (desc or "")[:80],
                "u": round(-abs(qty) if ln.get("ReturnToStock", False) else 0.0, 2),
                "r": round(-abs(_num(ln.get("LineTotal"))), 2),
                "c": round(-abs(qty) * unit_cost, 2),
            })

    out = [l for l in out if l["d"] >= FULL_START]
    out.sort(key=lambda x: x["d"])
    dts = [l["d"] for l in out]
    rng = [dts[0], dts[-1]] if dts else [None, None]
    return out, rng, products


# ------------------------------------------------------------------- LIVE: stock on hand
def pull_stock(c, products):
    bikes = {}
    for p in products:
        code = str(p.get("ProductCode", "") or "")
        if not code.upper().startswith(CURRENT_PREFIXES):
            continue
        grp = _gn(p, ["ProductGroup", "GroupName"]) or _gn(p, ["ProductGroup", "Name"])
        if isinstance(p.get("ProductGroup"), str):
            grp = p["ProductGroup"]
        if grp != "Bikes":
            continue
        bikes[code] = p.get("ProductDescription", "") or code
    stock = unleashed_all(c, "StockOnHand", {}, page_size=1000)
    qty = {}
    for s in stock:
        code = s.get("ProductCode", "")
        if code in bikes:
            q = s.get("AvailableQty")
            if q is None:
                q = s.get("QtyOnHand") or 0
            qty[code] = qty.get(code, 0) + int(_num(q))
    avail = []
    for code, name in bikes.items():
        q = qty.get(code, 0)
        model, size, colour = bl._variant_parts(name)
        avail.append({"v": name, "code": code, "m": model, "sz": size, "col": colour,
                      "q": q, "st": "In stock" if q > 0 else "Out of stock"})
    avail.sort(key=lambda x: -x["q"])
    return avail


# --------------------------------------------------------- LIVE: Campaign Monitor engagement
def cm_get(c, path):
    auth = base64.b64encode((c["cm_key"] + ":x").encode()).decode()
    req = urllib.request.Request(CM_BASE + path, headers={"Authorization": "Basic " + auth})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.load(r)


def pull_email(c):
    camps, page = [], 1
    while True:
        d = cm_get(c, "/clients/%s/campaigns.json?page=%d&pagesize=1000" % (c["cm_client"], page))
        camps.extend(d.get("Results", []))
        if page >= (d.get("NumberOfPages") or 1):
            break
        page += 1
    print("[live] Campaign Monitor: %d campaigns; pulling summaries…" % len(camps), flush=True)

    def summary(camp):
        try:
            s = cm_get(c, "/campaigns/%s/summary.json" % camp["CampaignID"])
        except Exception as e:  # noqa: BLE001 — one bad campaign never kills the run
            print("  ! summary failed for %s: %s" % (camp["Name"][:40], e), flush=True)
            s = {}
        return {
            "name": camp.get("Name", "")[:80],
            "date": (camp.get("SentDate") or "")[:10],
            "recipients": int(_num(s.get("Recipients") or camp.get("TotalRecipients"))),
            "uopens": int(_num(s.get("UniqueOpened"))),
            "uclicks": int(_num(s.get("Clicks"))),
            "topens": int(_num(s.get("TotalOpened"))),
            "unsub": int(_num(s.get("Unsubscribed"))),
            "bounced": int(_num(s.get("Bounced"))),
        }

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        rows = list(ex.map(summary, camps))
    rows.sort(key=lambda r: r["date"], reverse=True)
    tot = {
        "campaigns": len(rows),
        "sends": sum(r["recipients"] for r in rows),
        "opens": sum(r["topens"] for r in rows),
        "clicks": sum(r["uclicks"] for r in rows),
        "unique_opens": sum(r["uopens"] for r in rows),
    }
    return {
        "enabled": True, "mode": "live", "totals": tot, "campaigns": rows,
        "extracted_through": max((r["date"] for r in rows if r["date"]), default=None),
        "note": "Live Campaign Monitor pull — unique opens/clicks per campaign from the "
                "summary endpoint.",
    }


# --------------------------------------------------------- LIVE: Meta Ads via Windsor.ai
WINDSOR_URL = "https://connectors.windsor.ai/all"
WINDSOR_ACCOUNT = os.environ.get("WINDSOR_ACCOUNT", "facebook__465444904516684")
WINDSOR_PRESET = os.environ.get("WINDSOR_PRESET", "last_365d")
_META_FIELDS = ["date", "campaign", "adset_name", "ad_name", "spend", "impressions", "clicks",
                "link_clicks", "reach", "frequency", "actions_lead",
                "actions_landing_page_view"]


def pull_meta(c, previous):
    """Per-ad/day Meta rows via the Windsor `all` connector. Windsor caps a wide pull at ~12
    months, so each run fetches WINDSOR_PRESET and older rows are CARRIED FORWARD from the
    previous publication keyed by (date, campaign, adset, ad) — the honeytribe merge pattern.
    No key / any failure -> the previous publication's meta section (or a disabled stub)."""
    prev_meta = (previous or {}).get("meta_ads") or {}
    fallback = prev_meta if prev_meta.get("enabled") else {
        "enabled": False, "rows": [],
        "note": "Meta per-ad/day rows via Windsor.ai (spend, impressions, reach, link clicks, "
                "leads, frequency)."}
    if not c.get("w_key"):
        print("[meta] no Windsor key — carrying the previous section forward", flush=True)
        return fallback
    try:
        q = urllib.parse.urlencode({
            "api_key": c["w_key"], "date_preset": WINDSOR_PRESET,
            "fields": ",".join(_META_FIELDS), "select_accounts": WINDSOR_ACCOUNT})
        with urllib.request.urlopen(WINDSOR_URL + "?" + q, timeout=300) as r:
            data = json.load(r).get("data", [])
    except Exception as e:  # noqa: BLE001 — Windsor being down never kills the publish
        body = ""
        if hasattr(e, "read"):
            try:
                body = e.read().decode()[:160]
            except Exception:  # noqa: BLE001
                body = ""
        print("[meta] Windsor pull failed (%s %s) — carrying forward" % (e, body), flush=True)
        return fallback

    seen = {}
    for row in data:
        d = bl._date(row.get("date"))
        if not d:
            continue
        camp = str(row.get("campaign") or "").strip()
        adset = str(row.get("adset_name") or "").strip()
        ad = str(row.get("ad_name") or "").strip()
        seen[(d, camp, adset, ad)] = {
            "d": d, "camp": camp, "adset": adset, "ad": ad,
            "sp": round(_num(row.get("spend")), 2),
            "im": int(_num(row.get("impressions"))),
            "cl": int(_num(row.get("clicks"))),
            "lc": int(_num(row.get("link_clicks"))),
            "re": int(_num(row.get("reach"))),
            "fq": round(_num(row.get("frequency")), 4),
            "ld": int(_num(row.get("actions_lead"))),
            "lpv": int(_num(row.get("actions_landing_page_view"))),
        }
    if not seen:
        print("[meta] Windsor returned no rows — carrying forward", flush=True)
        return fallback
    fresh_min = min(k[0] for k in seen)
    kept = 0
    for old in prev_meta.get("rows", []):
        if old.get("d") and old["d"] < fresh_min:
            key = (old["d"], old.get("camp", ""), old.get("adset", ""), old.get("ad", ""))
            if key not in seen:
                seen[key] = old
                kept += 1
    rows = sorted(seen.values(), key=lambda x: (x["d"], x["camp"], x["ad"]))
    dates = [r2["d"] for r2 in rows]
    print("[meta] LIVE: %d ad-day rows %s .. %s (%d carried forward)"
          % (len(rows), dates[0], dates[-1], kept), flush=True)
    return {
        "enabled": True, "mode": "live",
        "account": WINDSOR_ACCOUNT,
        "rows": rows,
        "range": [dates[0], dates[-1]],
        "campaigns": sorted({r2["camp"] for r2 in rows if r2["camp"]}),
    }


# ------------------------------------------------------------------------------------ main
def main():
    import openpyxl

    c = creds()
    missing = [k for k in ("u_id", "u_key", "cm_key", "cm_client") if not c.get(k)]
    if missing:
        print("[!] missing credentials: %s — set env vars or keep the notebooks in context/"
              % ", ".join(missing))
        return 1

    previous = load_previous()
    sales, srange, products = pull_sales(c)
    avail = pull_stock(c, products)
    email = pull_email(c)
    meta_ads = pull_meta(c, previous)

    # CRM: live sheet -> dropped workbook -> previous publication (carry-forward).
    crm_mode = "live sheet"
    crm = pull_crm_sheet()
    if crm is None and os.path.exists(bl.XLSX_CRM):
        crm_mode = "workbook snapshot"
        print("[snapshot] CRM from %s" % os.path.basename(bl.XLSX_CRM), flush=True)
        wb_crm = openpyxl.load_workbook(bl.XLSX_CRM, read_only=True, data_only=True)
        crm = bl.build_crm(wb_crm)
        wb_crm.close()
    if crm is None and previous and previous.get("crm"):
        crm_mode = "carried forward"
        print("[carry] CRM from the previous publication", flush=True)
        crm = previous["crm"]
    if crm is None:
        crm = {"riders": [], "requests": [], "events": [], "range": [None, None]}
        crm_mode = "unavailable"

    # Production orders: live Lark -> dropped workbook -> previous publication.
    onway = pull_lark()
    onway_mode = "live Lark"
    stamp = None
    inv_fallback = None
    if os.path.exists(bl.XLSX_SALES):
        wb_sales = openpyxl.load_workbook(bl.XLSX_SALES, read_only=True, data_only=True)
        inv_fallback = bl.build_inventory(wb_sales)
        stamp = bl.build_meta_stamp(wb_sales)
        wb_sales.close()
    if onway is None and inv_fallback:
        onway = inv_fallback["onway"]
        onway_mode = "workbook snapshot"
    if onway is None and previous:
        onway = ((previous.get("inventory") or {}).get("onway")) or []
        onway_mode = "carried forward"
    if onway is None:
        onway = []
        onway_mode = "unavailable"
    if not avail and inv_fallback:      # a live-stock failure degrades to the workbook tab
        avail = inv_fallback["available"]
    if not avail and previous:
        avail = ((previous.get("inventory") or {}).get("available")) or []
    if not stamp:
        stamp = {"label": "Live pull", "tz": "Pacific/Auckland",
                 "refreshed": datetime.datetime.now(NZ_TZ).strftime("%Y-%m-%d %H:%M")}

    data = {
        "client": "MeloYelo",
        "tagline": "just mad about E-BIKES",
        "location": "New Zealand",
        "currency": "NZD",
        "generated_at": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "data_through": max(x for x in [srange[1], crm["range"][1]] if x),
        "fy_start_month": bl.FY_START_MONTH,
        "targets": {"fy_bikes": bl.ANNUAL_BIKE_TARGET},
        "source": {
            "sales": "Unleashed API (LIVE pull — invoices, credits, COGS at time of sale)",
            "crm": "CRM via %s" % crm_mode,
            "inventory": "Unleashed StockOnHand (LIVE) + production orders via %s" % onway_mode,
            "email": "Campaign Monitor API (LIVE per-campaign summaries)",
            "mode": "live",
        },
        "brand": {
            "mark": bl._b64(os.path.join(_HERE, "assets", "meloyelo-mark.svg"), "image/svg+xml"),
            "agora_logo": bl._b64(os.path.join(_HERE, "assets", "agora.png"), "image/png"),
        },
        "stamp": stamp,
        "sales": {"lines": sales, "range": srange},
        "crm": crm,
        "inventory": {"available": avail, "onway": onway},
        "email": email,
        "meta_ads": meta_ads,
        "google_ads": {"enabled": False, "rows": [],
                       "note": "Google Ads account 668-008-6591 via Windsor.ai (campaign/day: "
                               "cost, impressions, clicks, conversions)."},
        "ga4": {"enabled": False, "rows": [],
                "note": "GA4 via Windsor.ai (date/channel: sessions, engaged sessions, active "
                        "users, engagement rate)."},
    }

    body = json.dumps(data, separators=(",", ":"))
    if GCS_BUCKET:
        _gcs_bucket().blob(DATA_OBJECT).upload_from_string(
            body.encode("utf-8"), content_type="application/json")
        print("[meloyelo] published gs://%s/%s (LIVE)" % (GCS_BUCKET, DATA_OBJECT))
    else:
        os.makedirs(os.path.dirname(bl.OUT), exist_ok=True)
        with open(bl.OUT, "wb") as fh:  # BYTES on purpose — never text-mode JSON on Windows
            fh.write(body.encode("utf-8"))
        print("[meloyelo] wrote %s (LIVE)" % bl.OUT)
    print("   sales lines %6d   %s .. %s" % (len(sales), srange[0], srange[1]))
    print("   riders      %6d   requests %d   events %d   (CRM via %s, %s .. %s)"
          % (len(crm["riders"]), len(crm["requests"]), len(crm["events"]),
             crm_mode, crm["range"][0], crm["range"][1]))
    print("   stock       %6d variants (LIVE)   %d production orders (%s)"
          % (len(avail), len(onway), onway_mode))
    t = email["totals"]
    print("   email       %d campaigns LIVE — %s delivered, %s unique opens, %s clicks"
          % (t["campaigns"], t["sends"], t["unique_opens"], t["clicks"]))
    print("   size        %6d KB" % (len(body) // 1024))
    return 0


if __name__ == "__main__":
    sys.exit(main())
